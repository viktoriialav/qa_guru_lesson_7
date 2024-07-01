from openpyxl import load_workbook

workbook = load_workbook("tmp/file_example_XLSX_50.xlsx")
sheet = workbook.active
print(sheet.cell(row=2, column=3).value)
